import openpyxl as xl
from openpyxl.styles import Alignment
import datetime as dt
from math import ceil


ALPHA = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]


def build_individuals(dancers, feis_name, comp_name, judge_names):
    """(list of Dancer, str, str, list of str) -> NoneType
    Builds individual tabulated sheets for each dancer telling them their marks and the competition's marks. Saves
    these storage to to_print/<competition name>/<dancer number>.xlsx for all dancer numbers in the competition.
    """
    num_judges = len(dancers[0].scores)
    num_rounds = len(dancers[0].scores[0]) - 1

    wb = xl.load_workbook("outlines/" + str(num_judges) + "j" + str(num_rounds) + "r" + ".xlsx")
    ws = wb["Sheet1"]

    # General setup for all dancers.
    ws["A1"] = feis_name
    ws["G1"] = comp_name
    ws["M1"] = dt.datetime.now().strftime("%Y-%m-%d")
    for i in range(len(judge_names)):
        ws[ALPHA[2 + i * 4] + "4"] = judge_names[i]

    # Enter info for placed dancers
    # This is'nt proper, allow the user to choose how many dancers to place.
    end = int(ceil(len(dancers) / 2)) if len(dancers) > 5 else len(dancers)
    for j in range(end):
        # Enter place of current dancer.
        ws["A" + str(6 + j * 2)] = dancers[j].place
        ws["B" + str(6 + j * 2)] = dancers[j].dancer_num
        ws["B" + str(7 + j * 2)] = "Spencer McMurray"  # Get their name
        ws["N" + str(7 + j * 2)] = "BFOC"  # Get their school
        curr = ws["N" + str(7 + j * 2)]
        alignment = Alignment(horizontal="right")
        curr.alignment = alignment

        # Each judge.
        for k in range(len(dancers[j].scores)):
            # Each round for the current judge.
            for l in range(len(dancers[j].scores[k])):
                # Holy crap that's ugly... But its all about spacing and what not.
                ws[ALPHA[(2 + l + k * 4)] + str(6 + j * 2)] = dancers[j].scores[k][l]
        # Add the total at the end.
        total = 0
        for k in range(num_judges):
            total += ws[ALPHA[5 + k * 4] + str(6 + j * 2)].value
        ws["O" + str(6 + j * 2)] = total

    # Enter data for not placing dancers.
    for j in range(end, len(dancers)):
        # Enter place of current dancer.
        ws["A" + str(6 + end + j)] = dancers[j].place

        # Each judge.
        for k in range(len(dancers[j].scores)):
            # Each round for the current judge.
            for l in range(len(dancers[j].scores[k])):
                # Holy crap that's ugly... But its all about spacing and what not.
                ws[ALPHA[(2 + l + k * 4)] + str(6 + j + end)] = dancers[j].scores[k][l]
        total = 0
        for k in range(num_judges):
            total += ws[ALPHA[5 + k * 4] + str(6 + j + end)].value
        ws["O" + str(6 + end + j)] = total

    # Enter personalized info.
    for i in range(len(dancers)):
        ws["A3"] = dancers[i].place
        ws["B3"] = dancers[i].dancer_num
        ws["C2"] = "Spencer McMurray"  # Get their name
        ws["G2"] = "BFOC"  # Get their school
        for k in range(len(dancers[i].scores)):
            # Each round for the current judge.
            for l in range(len(dancers[i].scores[k])):
                # Holy crap that's ugly... But its all about spacing and what not.
                ws[ALPHA[(2 + l + k * 4)] + "3"] = dancers[i].scores[k][l]
        total = 0
        for k in range(num_judges):
            total += ws[ALPHA[5 + k * 4] + "3"].value
        ws["O3"] = total
        wb.save("to_print/" + comp_name + "/dancer#" +
                str(int(dancers[i].dancer_num)) + ".xlsx")


def build_placers(dancers, comp_name):
    """(list of Dancer, str) -> NoneType
    Builds the overall results for the competition. Saves to to_print/<comp name>/overall.xlsx
    """
    wb = xl.load_workbook("outlines/overall.xlsx")
    ws = wb["Sheet1"]
    ws["E1"] = comp_name

    # This is'nt proper, allow the user to choose how many dancers to place.
    end = int(ceil(len(dancers) / 2)) if len(dancers) > 5 else len(dancers)

    for i in range(end):
        ws["A" + str(i * 2 + 2)] = dancers[i].place
        ws["B" + str(i * 2 + 2)] = dancers[i].dancer_num
        ws["C" + str(i * 2 + 2)] = "Spencer McMurray"  # Get their name
        ws["C" + str(i * 2 + 3)] = "BFOC"  # Get their school.
    wb.save("to_print/" + comp_name + "/" + "overall" + ".xlsx")


def build_round_placers(dancers, comp_name, judge_names, judge_for_round):
    """(list of list of Dancer, str, list of str, bool) -> NoneType
    Builds the round medals for the competition. Saves to to_print/<comp name>/round_medals.xlsx
    """
    wb = xl.load_workbook("outlines/round.xlsx")
    ws = wb["Sheet1"]
    ws["A1"] = comp_name

    # Give each judge one round if judge_for_round, otherwise just fill in all with one judge.
    if judge_for_round:
        for judge in range(len(judge_names)):
            ws[ALPHA[2 + judge * 4] + "1"] = judge_names[judge]
    else:
        for judge in range(3):
            ws[ALPHA[2 + judge * 4] + "1"] = judge_names[0]

    # This is'nt proper, allow the user to choose how many dancers to place.
    end = int(ceil(len(dancers[0]) / 2)) if len(dancers[0]) > 5 else len(dancers[0])
    for i in range(len(dancers)):
        for j in range(end):
            ws[ALPHA[i * 4] + str(4 + j * 2)] = dancers[i][j].place
            ws[ALPHA[1 + i * 4] + str(4 + j * 2)] = dancers[i][j].dancer_num
            ws[ALPHA[2 + i * 4] + str(4 + j * 2)] = "Spencer McMurray"
            ws[ALPHA[2 + i * 4] + str(5 + j * 2)] = "BFOC"
    wb.save("to_print/" + comp_name + "/" + "round_medals" + ".xlsx")
